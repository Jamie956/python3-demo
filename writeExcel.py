# -*- coding: utf-8 -*-
from openpyxl import Workbook
import os

wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
# 插入下一行
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

ws.cell(3, 1, '111')
ws.cell(3, 2, '222')

# 如果不存在，创建output 文件夹
ouputExist = os.path.exists('./output')
# if ouputExist:
#     print('文件夹存在，删除输出目录 ./output')
#     os.removedirs('./output')

if bool(1-ouputExist):
    print('文件夹不存在, 创建输出目录 ./output')
    os.mkdir('./output')

# Save the file
wb.save("./output/target.xlsx")
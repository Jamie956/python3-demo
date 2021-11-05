# -*- coding: utf-8 -*-
import openpyxl
import sys
from datetime import datetime
import os


def simple_read():
    # 读excel
    wb = openpyxl.load_workbook('./output/target.xlsx')

    # 选择指定sheet
    # ws = wb['gov_Information']

    ws = wb.active
    rows = tuple(ws.rows)

    for row in rows:
        row_value = ''
        for cell in row:
            row_value += str(cell.value) + ' '
        print(row_value)

    '''
    min_col=4: 从第4列开始读
    min_row=2: 从第二行开始读
    max_row=2: 读到第二行结束
    '''
    row_datas = []
    for datas in ws.iter_rows(min_row=2, min_col=4, max_row=2):  # 只读取第二行
        for i in range(len(datas)):
            value = datas[i].value
            row_datas.append(value)

    '''
    min_row=3：从第三行开始读
    min_col=1：从第一列开始读
    max_col=1：读到第一列结束
    '''
    col_datas = []
    for datas in ws.iter_rows(min_row=3, min_col=1, max_col=1):
        for i in range(len(datas)):
            value = datas[i].value
            col_datas.append(value)


# 插入大量测试数据
def init_scroll_data():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(
        ["column1", "column2", "column3", "column4", "column5", "column6", "column7", "column8", "column9", "column10",
         "column11"])
    for i in range(10000):
        ws.append([i, 1, 2, 3, 4, 5, 6, 7, 8, 9, 0])
    wb.save("./output/target.xlsx")


'''
# 耗时（秒） 0.844773，read_only=False
wb = openpyxl.load_workbook('./output/target.xlsx',read_only=False)
# 耗时（秒） 0.528589，read_only=True
wb = openpyxl.load_workbook('./output/target.xlsx', read_only=True)
# 结论：read_only=True 可以加快读取速度

# excel 有10000 行数据，读100行，占用内存 0.00011444091796875 MB
# print(sys.getsizeof(ws.iter_rows(min_row=1, min_col=1, max_row=100)) / 1024 / 1024, 'MB')
# excel 有10000 行数据，读5000行，占用内存 0.00011444091796875 MB
# print(sys.getsizeof(ws.iter_rows(min_row=1, min_col=1, max_row=5000)) / 1024 / 1024, 'MB')
# excel 有10000 行数据，读全部行，占用内存 0.00011444091796875 MB
# print(sys.getsizeof(ws.rows) / 1024 / 1024, 'MB')
# 结论：iter_rows和rows 没有真实读取数据

# 读取真实数据 0.7629852294921875 MB
# print(sys.getsizeof(tuple(ws.rows)) / 1024 / 1024, 'MB')

'''


def scroll_read():
    start = datetime.now()
    wb = openpyxl.load_workbook('./output/target.xlsx', read_only=True)
    ws = wb.active

    # 读取起始行
    # row_cursor = 2
    # 每批数据的大小
    batch_size = 100

    headers = tuple()
    # for row in ws.iter_rows(min_row=1, max_row=1):
    #     headers = tuple(cell.value for cell in row)

    # print(sys.getsizeof(ws.values) / 1024 / 1024, 'MB')
    # print(sys.getsizeof(tuple(ws.values)) / 1024 / 1024, 'MB')
    flag = True
    items = []
    for row in ws.values:
        if flag:
            flag = False
            headers = row
            continue

        item = dict()
        for i in range(len(headers)):
            key = headers[i]
            value = row[i]
            item[key] = value
        items.append(item)

        if len(items) % 100 == 0:
            print()
    print()
    # headers = tuple(cell.value for cell in row)
    # for v in ws.values:
    #     print(v)
    # print()
    # total = ws.max_row
    # while total >= row_cursor:
    #     items = []
    #     rows = ws.iter_rows(min_row=row_cursor, max_row=batch_size)
    #     row_cursor += batch_size
    #     print('已读取行数', row_cursor - 1, '/', total)
    #     for row in rows:
    #         item = dict()
    #         values = tuple(cell.value for cell in row)
    #         # 第一行 header 和 row value 组成json，放入list
    #         for i in range(len(headers)):
    #             key = headers[i]
    #             value = values[i]
    #             item[key] = value
    #         items.append(item)
    #     print('已读取', row_cursor - 1, '/', total, '数据：', items)
    print('耗时（秒）', (datetime.now() - start).microseconds / 1000000, '读取')


def write_api():
    wb = openpyxl.Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = 42

    # Rows can also be appended
    # 插入下一行
    ws.append([1, 2, 3])

    # Python types will automatically be converted
    ws['A2'] = datetime.now()

    ws.cell(3, 1, '111')
    ws.cell(3, 2, '222')

    # 如果不存在，创建output 文件夹
    ouputExist = os.path.exists('./output')
    if bool(1 - ouputExist):
        print('文件夹不存在, 创建输出目录 ./output')
        os.mkdir('./output')

    # Save the file
    wb.save("./output/target.xlsx")


if __name__ == '__main__':
    # simple_read()
    # init_scroll_data()
    scroll_read()
    # write_api()
    print()

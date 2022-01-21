# -*- coding: utf-8 -*-
import openpyxl
from datetime import datetime
import re

'''
===================== api ===================== 
'''


def simple_read():
    # 读excel
    wb = openpyxl.load_workbook('./output/target.xlsx')
    # 选择指定sheet
    # ws = wb['gov_Information']

    ws = wb.active
    rows = tuple(ws.rows)

    for row in rows:
        row_value = ''
        for cell in row:
            row_value += str(cell.value) + ' '
        print(row_value)


def iter_read():
    wb = openpyxl.load_workbook('./output/target.xlsx')
    ws = wb.active

    '''
    min_col=4: 从第4列开始读
    min_row=2: 从第二行开始读
    max_row=2: 读到第二行结束
    '''
    row_datas = []
    for datas in ws.iter_rows(min_row=2, min_col=4, max_row=2):  # 只读取第二行
        for i in range(len(datas)):
            value = datas[i].value
            row_datas.append(value)

    print(row_datas)


def iter_read2():
    wb = openpyxl.load_workbook('./output/target.xlsx')
    ws = wb.active
    '''
    min_row=3：从第三行开始读
    min_col=1：从第一列开始读
    max_col=1：读到第一列结束
    '''
    col_datas = []
    for datas in ws.iter_rows(min_row=3, min_col=1, max_col=1):
        for i in range(len(datas)):
            value = datas[i].value
            col_datas.append(value)

    print(col_datas)


def write_api():
    workbook = openpyxl.Workbook()
    # grab the active worksheet
    default_worksheet = workbook.active
    # Data can be assigned directly to cells
    default_worksheet['A1'] = 42
    # Rows can also be appended
    default_worksheet.append([1, 2, 3])
    # Python types will automatically be converted
    default_worksheet['A2'] = datetime.now()
    # 指定坐标插入
    default_worksheet.cell(3, 1, 'insert31')
    default_worksheet.cell(3, 2, 'insert32')
    # Save the file
    workbook.save('./output/target.xlsx')


def create_sheet_api():
    new_sheet_name = 'sec-sheet'
    workbook = openpyxl.Workbook()
    workbook.remove(workbook['Sheet'])
    workbook.create_sheet(new_sheet_name)
    worksheet = workbook[new_sheet_name]
    worksheet.append([1, 2, 3])
    worksheet.append([4, 5, 6])
    workbook.save('./output/target.xlsx')


'''
===================== util ===================== 
'''


def read_as_json(config):
    print('read excel')
    wb = openpyxl.load_workbook(config['file_path'])
    ws = wb.active

    headers = []
    for items in ws.iter_rows(min_row=1, max_row=1):
        for i in range(len(items)):
            value = items[i].value
            headers.append(value)

    data = []
    for items in ws.iter_rows(min_row=2):
        row = {}
        for i in range(len(headers)):
            header = headers[i]
            value = items[i].value
            row[header] = value
        data.append(row)
    return tuple(data)


def batch_create_sheet(wb, sheets):
    for sheet in sheets:
        wb.create_sheet(sheet)


def json_arr_to_excel(worksheet, json_array, fields):
    for json_data in json_array:
        append_row = []
        for j, field in enumerate(fields):
            if field in json_data.keys():
                json_field_value = str(json_data[field])
                # 去除excel非法字符，否在抛出异常 openpyxl.utils.exceptions.IllegalCharacterError
                json_field_value = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]').sub(r'', json_field_value)
                append_row.append(json_field_value)
            else:
                # 字段不存在，写入空
                append_row.append('')
        worksheet.append(append_row)


'''
    data_config = [
        {
            'sheet': 'mysheet1',
            'headers': ['name', 'age'],
            'data': [{'name': '12', 'age': 11, 'ad': 'as'},
                          {'name': 'to', 'age': 334, 'ad': 'ass'}]
        },
        {
            'sheet': 'mysheet2',
            'headers': ['name', 'age', 'ad'],
            'data': [{'name': '12', 'age': 11, 'ad': 'as'},
                           {'name': 'to', 'age': 334, 'ad': 'ass'}]
        }
    ]
'''


def batch_write2excel_by_config(output_path, data_config):
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    for cf in data_config:
        sheet_name = cf['sheet']
        headers = cf['headers']
        json_array = cf['data']
        batch_create_sheet(workbook, [sheet_name])

        worksheet = workbook[sheet_name]
        worksheet.append(headers)
        json_arr_to_excel(worksheet, json_array, headers)

    workbook.save(output_path)


if __name__ == '__main__':
    # config = {
    #     'file_path': './output/target.xlsx',
    #     'sheet': ''
    # }
    # print(read_as_json(config))

    data_config = [
        {
            'sheet': 'mysheet1',
            'headers': ['name', 'age'],
            'data': [{'name': '12', 'age': 11, 'ad': 'as'},
                           {'name': 'to', 'age': 334, 'ad': 'ass'}]
        },
        {
            'sheet': 'mysheet2',
            'headers': ['name', 'age', 'ad'],
            'data': [{'name': 'cat', 'age': 11, 'ad': 'as'},
                           {'name': 'to', 'age': 334, 'ad': 'ass'}]
        }
    ]

    batch_write2excel_by_config('./output/target.xlsx', data_config)

    print()
